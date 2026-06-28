import openpyxl

sys_wb = openpyxl.load_workbook(r'C:\Users\jhoeb\Downloads\Lucerne_2026-05-24_2026-06-08_2026-06-15_payroll (8).xlsx', data_only=True)
ref_wb = openpyxl.load_workbook(r'C:\Users\jhoeb\Downloads\11_Alphalist May 24 - June 8, 2026_Agnes.xlsx', data_only=True)
sys_ws = sys_wb.active
ref_ws = ref_wb['Payroll']

# System: key = (acct, subclient, salary, total_allow)
sys_data = {}
for i in range(4, sys_ws.max_row + 1):
    acct = sys_ws.cell(i, 7).value
    if not acct:
        continue
    acct = str(acct).strip()
    subclient = str(sys_ws.cell(i, 8).value or '').strip().upper()
    salary = float(sys_ws.cell(i, 10).value or 0)
    total_allow = sum(float(v or 0) for v in [sys_ws.cell(i, 12).value, sys_ws.cell(i, 13).value, sys_ws.cell(i, 14).value, sys_ws.cell(i, 16).value])
    gross = round(float(sys_ws.cell(i, 67).value or 0), 2)
    name = sys_ws.cell(i, 3).value or ''
    key = (acct, subclient, salary, total_allow)
    emp_info = f'Emp Name: {name}\nAccount Number: {acct}\nSubclient: {subclient}\nSalary: {salary}\nTotal Allowance: {total_allow}'
    sys_data[key] = {'gross': gross, 'info': emp_info}

# Reference: key = (acct, subclient, salary, total_allow)
ref_data = {}
for i in range(11, ref_ws.max_row + 1):
    acct = ref_ws.cell(i, 3).value
    if not acct:
        continue
    acct = str(acct).strip()
    subclient = str(ref_ws.cell(i, 2).value or '').strip().upper()
    salary = float(ref_ws.cell(i, 9).value or 0)
    total_allow = sum(float(v or 0) for v in [ref_ws.cell(i, 10).value, ref_ws.cell(i, 11).value, ref_ws.cell(i, 12).value, ref_ws.cell(i, 13).value])
    gross = round(float(ref_ws.cell(i, 66).value or 0), 2)
    name = ref_ws.cell(i, 4).value or ''
    key = (acct, subclient, salary, total_allow)
    emp_info = f'Emp Name: {name}\nAccount Number: {acct}\nSubclient: {subclient}\nSalary: {salary}\nTotal Allowance: {total_allow}'
    ref_data[key] = {'gross': gross, 'info': emp_info}

# Compare
differences, matches, not_in_ref, not_in_sys = [], [], [], []

for key, s in sys_data.items():
    if key in ref_data:
        r = ref_data[key]
        diff = round(s['gross'] - r['gross'], 2)
        row = [s['info'], r['info'], s['gross'], r['gross'], diff]
        if abs(diff) > 0.01:
            differences.append(row)
        else:
            matches.append(row)
    else:
        not_in_ref.append([s['info'], '', s['gross'], '', ''])

for key, r in ref_data.items():
    if key not in sys_data:
        not_in_sys.append(['', r['info'], '', r['gross'], ''])

# Write output
out_wb = openpyxl.Workbook()
headers = ['System File Emp Info', 'Reference File Emp Info', 'System Gross Pay', 'Reference Gross Pay', 'Difference']

def write_sheet(ws, title, data):
    ws.title = title
    ws.append(headers)
    for row in data:
        ws.append(row)

write_sheet(out_wb.active, 'Differences', differences)
write_sheet(out_wb.create_sheet(), 'Matches', matches)
write_sheet(out_wb.create_sheet(), 'Not In Reference', not_in_ref)
write_sheet(out_wb.create_sheet(), 'Not In System', not_in_sys)

out_path = r'C:\Users\jhoeb\Downloads\Payroll_Comparison_Result.xlsx'
out_wb.save(out_path)
print(f'Done! Saved to: {out_path}')
print(f'Differences (>0.01): {len(differences)}')
print(f'Matches: {len(matches)}')
print(f'Not in Reference: {len(not_in_ref)}')
print(f'Not in System: {len(not_in_sys)}')
print(f'System entries: {len(sys_data)}')
print(f'Reference entries: {len(ref_data)}')
